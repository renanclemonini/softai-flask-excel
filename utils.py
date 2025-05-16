import pandas as pd
import os
import re
from openpyxl import load_workbook, Workbook
from uuid import uuid4
from openpyxl.styles import Alignment, Font
from datetime import time
import xlrd
from datetime import datetime
from functools import wraps
import threading
import time

def limpar_pasta_input():
    def agendador():
        while True:
            pasta = 'output'
            for arquivo in os.listdir(pasta):
                caminho = os.path.join(pasta, arquivo)
                try:
                    if os.path.isfile(caminho):
                        os.remove(caminho)
                except Exception as e:
                    print(f"Erro ao remover {arquivo}: {e}")
            time.sleep(3600)  # 3600 segundos = 1 hora

    t = threading.Thread(target=agendador, daemon=True)
    t.start() 


def medidor_tempo(bench_mark=False):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            if bench_mark:
                inicio_tempo = datetime.now()
                print(f"[{func.__name__}] \nInício: {inicio_tempo.strftime('%H:%M:%S')}")

            resultado = func(*args, **kwargs)

            if bench_mark:
                fim_tempo = datetime.now()
                diferenca = fim_tempo - inicio_tempo
                diferenca_ms = diferenca.total_seconds() * 1000
                print(f"Fim: {fim_tempo.strftime('%H:%M:%S.%f')}")
                if diferenca_ms >= 1000:
                    diferenca_s = diferenca_ms / 1000
                    print(f"Diferença: {diferenca_s:.2f} segundos \n[{func.__name__}]")
                else:
                    print(f"Diferença: {diferenca_ms:.2f} ms \n[{func.__name__}]")

            return resultado
        return wrapper
    return decorator

def processar_excel(arquivo_entrada):
    try:
        # Lê o arquivo Excel
        df = pd.read_excel(arquivo_entrada)

        # Colunas obrigatórias
        colunas_obrigatorias = ['Primeiro nome', 'Sobrenome', 'Telefone', 'Etiquetas']

        # Verifica se todas as colunas existem
        for col in colunas_obrigatorias:
            print(type(col))
            if col not in df.columns:
                raise ValueError(f"A coluna obrigatória '{col}' não foi encontrada.")

        # Remove linhas completamente vazias
        df = df.dropna(how='all')

        # Remove colunas completamente vazias
        df = df.dropna(axis=1, how='all')

        nome_arquivo = arquivo_entrada.split('.')[0]
        arquivo_saida = f'{nome_arquivo}_ok.xlsx'
        # Salva em novo arquivo Excel
        df.to_excel(arquivo_saida, index=False)
        print(f"Arquivo processado com sucesso: {arquivo_saida}")
        
    except Exception as e:
        print(f"Erro ao processar o arquivo: {e}")


def processar_excel_2(arquivo_entrada):
    try:
        # Lê o arquivo Excel
        df = pd.read_excel(arquivo_entrada)
        df['Telefone'] = df['Telefone'].astype(str)

        # Colunas obrigatórias
        colunas_obrigatorias = ['Primeiro nome', 'Sobrenome', 'Telefone', 'Etiquetas']

        # Verifica se todas as colunas existem
        for col in colunas_obrigatorias:
            if col not in df.columns:
                raise ValueError(f"A coluna obrigatória '{col}' não foi encontrada.")

        # Remove linhas e colunas completamente vazias
        df = df.dropna(how='all')
        df = df.dropna(axis=1, how='all')

        # Resetar índice para evitar problemas ao excluir linhas
        df = df.reset_index(drop=True)

        linhas_para_remover = []

        for i, row in df.iterrows():
            # --- PROCESSA 'Primeiro nome' + 'Sobrenome'
            primeiro_nome = str(row['Primeiro nome']).strip()
            partes = primeiro_nome.split()
            if len(partes) > 1:
                # Pega o primeiro como primeiro nome
                df.at[i, 'Primeiro nome'] = partes[0]
                # Junta os demais e adiciona ao início do sobrenome
                sobrenome = str(row['Sobrenome']).strip().lower()
                df.at[i, 'Sobrenome'] = ' '.join(partes[1:]).title() + ' ' + sobrenome.title()

            # --- PROCESSA 'Telefone'
            telefone = str(row['Telefone'])
            telefone_numerico = re.sub(r'\D', '', telefone)  # remove tudo que não for número
            if len(telefone_numerico) > 13:
                linhas_para_remover.append(i)
            else:
                df.at[i, 'Telefone'] = telefone_numerico

            # --- PROCESSA 'Etiquetas'
            etiquetas = str(row['Etiquetas']).strip()
            if etiquetas and etiquetas.lower() != 'nan':
                df.at[i, 'Etiquetas'] = etiquetas + ', etiqueta_valida'
            else:
                df.at[i, 'Etiquetas'] = 'etiqueta_valida'

        # Remove linhas com telefone inválido
        df = df.drop(index=linhas_para_remover).reset_index(drop=True)

        # Gera nome do arquivo de saída
        nome_arquivo = arquivo_entrada.rsplit('.', 1)[0]
        arquivo_saida = f'{nome_arquivo}_planilha_ok.xlsx'

        # Salva em novo arquivo Excel
        df.to_excel(arquivo_saida, index=False)
        print(f"Arquivo processado com sucesso: {arquivo_saida}")

    except Exception as e:
        print(f"Erro ao processar o arquivo: {e}")


def aplicar_formatacao_excel(caminho_arquivo):
    wb = load_workbook(caminho_arquivo)
    ws = wb.active

    for cell in ws[1]:  # Primeira linha = cabeçalho
        cell.font = Font(bold=False)  # Remove negrito
        cell.alignment = Alignment(horizontal='left')  # Alinha à esquerda

    wb.save(caminho_arquivo)


def processar_excel_3(arquivo_entrada):
    try:
        df = pd.read_excel(f'{arquivo_entrada}.xlsx')

        # Padroniza os nomes das colunas
        df.columns = df.columns.str.strip().str.replace(r'\s+', ' ', regex=True).str.lower()
        # list_df = list(df.columns)
        print("Colunas encontradas:", list(df.columns))

        # Verifica colunas obrigatórias (em minúsculo)
        colunas_obrigatorias = ['primeiro nome', 'sobrenome', 'telefone', 'etiquetas']
        for col in colunas_obrigatorias:
            if col not in df.columns:
                raise ValueError(f"A coluna obrigatória '{col}' não foi encontrada.")

        df['telefone'] = df['telefone'].astype(str)

        # Remove linhas e colunas completamente vazias
        df = df.dropna(how='all').dropna(axis=1, how='all').reset_index(drop=True)

        linhas_para_remover = []

        for i, row in df.iterrows():
            # Trata nomes
            primeiro_nome = str(row['primeiro nome']).strip().lower()
            sobrenome = str(row.get('sobrenome', '')).strip().lower()
            partes = primeiro_nome.split()

            if len(partes) > 1:
                df.at[i, 'primeiro nome'] = partes[0].title()
                sobrenome_completo = ' '.join(partes[1:] + [sobrenome])
                df.at[i, 'sobrenome'] = sobrenome_completo.title()
            else:
                df.at[i, 'primeiro nome'] = primeiro_nome.title()
                df.at[i, 'sobrenome'] = sobrenome.title()

            # Trata telefone
            telefone = re.sub(r'\D', '', str(row['telefone']))
            if len(telefone) > 13:
                linhas_para_remover.append(i)
            else:
                df.at[i, 'telefone'] = telefone

            # Trata etiquetas
            etiqueta = str(row['etiquetas']).strip()
            if etiqueta and etiqueta.lower() != 'nan':
                df.at[i, 'etiquetas'] = f"{etiqueta}, etiqueta_valida"
            else:
                df.at[i, 'etiquetas'] = 'etiqueta_valida'

        df = df.drop(index=linhas_para_remover).reset_index(drop=True)

        arquivo_saida = f'{arquivo_entrada}_planilha_ok.xlsx'
        df.to_excel(arquivo_saida, index=False)

        aplicar_formatacao_excel(arquivo_saida)

        print(f"Arquivo processado com sucesso: {arquivo_saida}")

    except Exception as e:
        print(f"Erro ao processar o arquivo: {e}")


def processar_excel_4(arquivo_entrada):
    try:
        #  Lê o Excel
        df = pd.read_excel(f'{arquivo_entrada}.xlsx')

        # Garante que haverá pelo menos 4 colunas
        if df.shape[1] < 4:
            raise ValueError("O arquivo deve conter pelo menos 4 colunas.")

        # Força o uso das 4 primeiras colunas e renomeia
        df = df.iloc[:, :4]
        df.columns = ['primeiro nome', 'sobrenome', 'telefone', 'etiquetas']

        print("Colunas padronizadas:", list(df.columns))

        df['telefone'] = df['telefone'].astype(str)

        # Remove linhas e colunas completamente vazias
        df = df.dropna(how='all').dropna(axis=1, how='all').reset_index(drop=True)

        linhas_para_remover = []

        for i, row in df.iterrows():
            # Trata 'Primeiro nome' e 'Sobrenome'
            primeiro_nome = str(row['primeiro nome']).strip().lower()
            sobrenome = str(row.get('sobrenome', '')).strip().lower() or ''
            partes = primeiro_nome.split()

            if len(partes) > 1:
                df.at[i, 'primeiro nome'] = partes[0].title()
                sobrenome_completo = ' '.join(partes[1:] + [sobrenome]) if sobrenome else ' '.join(partes[1:])
                df.at[i, 'sobrenome'] = sobrenome_completo.title()
            else:
                df.at[i, 'primeiro nome'] = primeiro_nome.title()
                df.at[i, 'sobrenome'] = sobrenome.title() if sobrenome else ''

            # Trata telefone
            telefone = re.sub(r'\D', '', str(row['telefone']))
            if len(telefone) > 13:
                linhas_para_remover.append(i)
            else:
                df.at[i, 'telefone'] = telefone

            # Trata etiquetas
            etiqueta = str(row['etiquetas']).strip()
            if etiqueta and etiqueta.lower() != 'nan':
                df.at[i, 'etiquetas'] = f"{etiqueta}, etiqueta_valida"
            else:
                df.at[i, 'etiquetas'] = 'etiqueta_valida'

        # Remove linhas com telefone inválido
        df = df.drop(index=linhas_para_remover).reset_index(drop=True)

        # Mantém apenas as colunas necessárias, na ordem correta
        df = df[['primeiro nome', 'sobrenome', 'telefone', 'etiquetas']]
        df.columns = ['Primeiro nome', 'Sobrenome', 'Telefone', 'Etiquetas']


        # Exporta o arquivo
        arquivo_saida = f'{arquivo_entrada}__ok.xlsx'
        df.to_excel(arquivo_saida, index=False)

        aplicar_formatacao_excel(arquivo_saida)

        print(f"Arquivo processado com sucesso: {arquivo_saida}")

    except Exception as e:
        print(f"Erro ao processar o arquivo: {e}")


def processar_excel_oficial(arquivo_entrada):
    try:
        with open(arquivo_entrada, 'rb') as file:
            # df = pd.read_excel(f'{arquivo_entrada}.xlsx')
            df = pd.read_excel(file)
            # Normaliza os nomes das colunas
            df.columns = df.columns.str.strip().str.replace(r'\s+', ' ', regex=True).str.lower()
            print("Colunas encontradas:", list(df.columns))

            # Garante que há pelo menos 3 colunas úteis
            if df.shape[1] < 3:
                raise ValueError("O arquivo deve conter pelo menos 3 colunas com dados relevantes.")

            # Cria um novo DataFrame padronizado
            novo_df = pd.DataFrame(columns=['primeiro nome', 'sobrenome', 'telefone', 'etiquetas'])

            # --- Trata nome e sobrenome
            if 'nome' in df.columns:
                nomes = df['nome'].astype(str).str.strip().str.split()
                novo_df['primeiro nome'] = nomes.str[0].fillna('').str.title()
                novo_df['sobrenome'] = nomes.str[1:].apply(lambda x: ' '.join(x)).fillna('').str.title()
            elif 'primeiro nome' in df.columns and 'sobrenome' in df.columns:
                novo_df['primeiro nome'] = df['primeiro nome'].astype(str).str.strip().str.title()
                novo_df['sobrenome'] = df['sobrenome'].astype(str).str.strip().str.title()
            else:
                raise ValueError("Colunas 'nome' ou 'primeiro nome' + 'sobrenome' são obrigatórias.")

            # --- Trata telefone
            if 'telefone' in df.columns:
                novo_df['telefone'] = df['telefone'].astype(str).apply(lambda t: re.sub(r'\D', '', t))
            # elif 'contato' in df.columns:
            #     novo_df['telefone'] = df['contato'].astype(str).apply(lambda t: re.sub(r'\D', '', t))
            else:
                raise ValueError("Coluna 'telefone' ou 'contato' é obrigatória.")

            # --- Trata etiquetas
            if 'etiquetas' in df.columns:
                novo_df['etiquetas'] = df['etiquetas'].astype(str).apply(
                    lambda e: f"{e.strip()}, etiqueta_valida" if e.strip().lower() != 'nan' else 'etiqueta_valida'
                )
            elif 'etiqueta' in df.columns:
                novo_df['etiquetas'] = df['etiqueta'].astype(str).apply(
                    lambda e: f"{e.strip()}, etiqueta_valida" if e.strip().lower() != 'nan' else 'etiqueta_valida'
                )
            else:
                novo_df['etiquetas'] = ''

            # Remove linhas com telefone inválido
            novo_df = novo_df[novo_df['telefone'].str.len() > 13].reset_index(drop=True)

            # Renomeia as colunas com inicial maiúscula
            novo_df.columns = ['Primeiro nome', 'Sobrenome', 'Telefone', 'Etiquetas']

            # Salva no arquivo de saída
            arquivo_saida = f'{arquivo_entrada}_{uuid4()}.xlsx'
            novo_df.to_excel(arquivo_saida, index=False)

            # Aplica formatação opcional
            aplicar_formatacao_excel(arquivo_saida)

            print(f"Arquivo processado com sucesso: {arquivo_saida}")
            return arquivo_saida

    except Exception as e:
        print(f"Erro ao processar o arquivo: {e}")


ALLOWED_EXTENSIONS = set(['xls', 'xlsx'])

def allowed_file(filename: str):
    return '.' in filename and \
        filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def processar_excel_oficial2(arquivo_entrada: str):
    with open(arquivo_entrada, 'rb') as f:
        wb = load_workbook(f)
        ws = wb.active
        headers = [str(cell.value).strip().lower() if cell.value else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        print("Colunas encontradas:", headers)

        idx = {h: i for i, h in enumerate(headers)}
        novo_dados = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            primeiro_nome = ''
            sobrenome = ''
            telefone = ''
            etiquetas = 'etiqueta_valida'

            # Nome completo ou separado
            if 'nome' in idx:
                nome = str(row[idx['nome']] or '').strip().split()
                primeiro_nome = nome[0].title() if nome else ''
                sobrenome = ' '.join(nome[1:]).title() if len(nome) > 1 else ''
            else:
                if 'primeiro nome' in idx:
                    primeiro_nome = str(row[idx['primeiro nome']] or '').strip().title()
                if 'sobrenome' in idx:
                    sobrenome = str(row[idx['sobrenome']] or '').strip().title()

            # Telefone
            for col in ['telefone', 'contato', 'celular']:
                if col in idx:
                    telefone = str(row[idx[col]] or '')
                    telefone = re.sub(r'\D', '', telefone)
                    break

            # Etiquetas
            for col in ['etiquetas', 'etiqueta', 'tag']:
                if col in idx:
                    etiqueta_padrao = 'NomeConfirmado'
                    val = str(row[idx[col]] or '').strip()
                    etiquetas = f'{val}, {etiqueta_padrao}' if val.lower() != 'nan' and val else etiqueta_padrao
                    break

            # Adiciona se telefone parecer válido
            if telefone and len(telefone) >= 10:
                novo_dados.append([primeiro_nome, sobrenome, telefone, etiquetas])

        # Criação do novo Excel
        wb_novo = Workbook()
        ws_novo = wb_novo.active
        ws_novo.append(['Primeiro nome', 'Sobrenome', 'Telefone', 'Etiquetas'])
        for linha in novo_dados:
            ws_novo.append(linha)

        nome_base = os.path.splitext(os.path.basename(arquivo_entrada))[0]
        nome_saida = f'{nome_base}_{uuid4()}.xlsx'
        caminho_saida = os.path.join('output', nome_saida)
        os.makedirs('output', exist_ok=True)
        wb_novo.save(caminho_saida)

        print(f"Arquivo processado com sucesso: {nome_saida}")
        return nome_saida
    
@medidor_tempo(True)
def processar_excel_oficial3(arquivo_entrada: str):
    with open(arquivo_entrada, 'rb') as f:
        wb = load_workbook(f)
        ws = wb.active
        linhas_originais = ws.max_row
        colunas_originais = ws.max_column
        headers = [str(cell.value).strip().lower() if cell.value else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        print(f"({len(headers)}) Colunas encontradas: {headers}")

        idx = {h: i for i, h in enumerate(headers)}
        novo_dados = []

        padrao_3_colunas = set(['telefone', 'nome', 'etiquetas']).issubset(set(headers))
        padrao_4_colunas = set(['primeiro nome', 'sobrenome', 'telefone', 'etiquetas']).issubset(set(headers))
        linhas_em_branco = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(cell is None or str(cell).strip() == '' for cell in row):
                linhas_em_branco += 1
                continue
            primeiro_nome = ''
            sobrenome = ''
            telefone = ''
            etiquetas = 'etiqueta_valida'

            if padrao_3_colunas:
                nome = str(row[idx['nome']] or '').strip()
                partes = nome.split()
                primeiro_nome = partes[0].title() if partes else ''
                sobrenome = ' '.join(partes[1:]).title() if len(partes) > 1 else ''

            elif padrao_4_colunas:
                primeiro = str(row[idx['primeiro nome']] or '').strip()
                sobrenome_original = str(row[idx['sobrenome']] or '').strip()

                partes = primeiro.split()
                primeiro_nome = partes[0].title() if partes else ''
                sobrenome_splitado = ' '.join(partes[1:]).title() if len(partes) > 1 else ''
                sobrenome = f"{sobrenome_splitado} {sobrenome_original}".strip().title()

            else:
                print("Formato de planilha não reconhecido.")
                continue

            # Telefone
            for col in ['telefone', 'contato', 'celular']:
                if col in idx:
                    telefone = str(row[idx[col]] or '')
                    telefone = re.sub(r'\D', '', telefone)
                    while len(telefone) > 13:
                        telefone = telefone[:4] + telefone[5:]
                    break

            # Etiquetas
            for col in ['etiquetas', 'etiqueta', 'tag']:
                if col in idx:
                    etiqueta_padrao = 'NomeConfirmado'
                    val = str(row[idx[col]] or '').strip()
                    etiquetas = f'{val}, {etiqueta_padrao}' if val.lower() != 'nan' and val else etiqueta_padrao
                    break

            if telefone and len(telefone) >= 10:
                novo_dados.append([primeiro_nome, sobrenome, telefone, etiquetas])
        colunas_em_branco = 0
        for col_idx in range(colunas_originais):
            coluna_vazia = True
            for row in ws.iter_rows(min_row=2, min_col=col_idx + 1, max_col=col_idx + 1, values_only=True):
                if row[0] is not None and str(row[0]).strip() != '':
                    coluna_vazia = False
                    break
            if coluna_vazia:
                colunas_em_branco += 1
        # Criação do novo Excel
        wb_novo = Workbook()
        ws_novo = wb_novo.active
        ws_novo.append(['Primeiro nome', 'Sobrenome', 'Telefone', 'Etiquetas'])
        for linha in novo_dados:
            ws_novo.append(linha)

        nome_base = os.path.splitext(os.path.basename(arquivo_entrada))[0]
        nome_saida = f'{nome_base}_{uuid4()}.xlsx'
        caminho_saida = os.path.join('output', nome_saida)
        os.makedirs('output', exist_ok=True)
        wb_novo.save(caminho_saida)
        response = {
            'arquivo_processado': nome_saida,
            'linhas_originais': linhas_originais,
            'colunas_originais': colunas_originais,
            'colunas_encontradas': headers,
            'linhas_novo': len(novo_dados),
            'linhas_em_branco': linhas_em_branco,
            'colunas_em_branco': colunas_em_branco
        }

        print(f"Arquivo original: {linhas_originais} linhas x {colunas_originais} colunas")
        print(f"Novo arquivo: {len(novo_dados) + 1} linhas x 4 colunas") 
        print(f"Arquivo processado com sucesso: {nome_saida}")
        return response


def converter_xls_para_xlsx(caminho_xls: str) -> str:
    if not caminho_xls.endswith('.xls'):
        return caminho_xls  # Se não for .xls, retorna o caminho original

    # Abre o arquivo .xls com xlrd
    workbook_xls = xlrd.open_workbook(caminho_xls)
    sheet_xls = workbook_xls.sheet_by_index(0)  # lê apenas a primeira aba

    # Cria um novo workbook .xlsx com openpyxl
    workbook_xlsx = Workbook()
    sheet_xlsx = workbook_xlsx.active
    sheet_xlsx.title = sheet_xls.name

    # Copia célula por célula
    for row in range(sheet_xls.nrows):
        for col in range(sheet_xls.ncols):
            value = sheet_xls.cell_value(row, col)
            sheet_xlsx.cell(row=row+1, column=col+1).value = value

    novo_caminho = caminho_xls + 'x'  # Ex: input/arquivo.xls → input/arquivo.xlsx
    workbook_xlsx.save(novo_caminho)

    return novo_caminho


